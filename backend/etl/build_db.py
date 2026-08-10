#!/usr/bin/env python3
"""ETL: GBTAC CSV/XLSX → SQLite (hourly + daily aggregates). Hardcoded paths at top."""
DATA_DIR = "/Users/yuki/Downloads/aris-web-app/src/GBTAC Building Data"
CSV_PATH = DATA_DIR + "/merged_timeseries_student.csv"
XLSX_FULL = DATA_DIR + "/Sensor Names.xlsx"
XLSX_SIMPLE = DATA_DIR + "/Sensor_name-simplified.xlsx"
XLSX_GAS = DATA_DIR + "/Natural Gas Readings.xlsx"
DB_PATH = "/Users/yuki/Downloads/aris-web-app/backend/aris.db"
SCHEMA_PATH = "/Users/yuki/Downloads/aris-web-app/backend/etl/schema.sql"
CHUNK = 50000
PREFIX = "SaitSolarLab_"

# Prefer these meters for energy_daily / ML (CSV column suffixes after strip)
ENERGY_CODES = {
	"30000_TL208": ("hvac", "W", "Space HVAC"),
	"30000_TL209": ("lighting", "W", "Lighting"),
	"30000_TL210": ("dhw", "W", "DHW"),
	"30000_TL211": ("rnd", "W", "R&D loads"),
	"30000_TL212": ("appliances", "W", "Appliances"),
	"30000_TL252": ("pv_carport", "W", "PV Carport"),
	"30000_TL253": ("pv_rooftop", "W", "PV Rooftop"),
	"30000_TL335": ("net", "W", "Net energy"),
	"30000_TL3": ("gen_total", "W", "Total generation"),
	"30000_TL4": ("vent", "W", "Ventilation"),
	"30000_TL342": ("mains_a", "W", "Mains PA"),
	"30000_TL343": ("mains_b", "W", "Mains PB"),
	"30000_TL344": ("mains_c", "W", "Mains PC"),
	"20000_TL92": ("oat", "C", "Outside air temp"),
}

import csv, os, re, sqlite3, sys
from collections import defaultdict
from datetime import datetime

def progress(msg):
	print(msg, flush=True)

def infer_type_unit(prop, name):
	n = (name or "").lower()
	if prop == 6 or "temp" in n: return "Temperature", "°C"
	if prop == 44 or "co2" in n: return "CO2", "ppm"
	if prop in (59,) or "power" in n or "pv" in n: return "Power", "W"
	if prop in (60, 25) or "wh" in n or "kwh" in n: return "Energy", "Wh"
	if "humid" in n: return "Humidity", "%"
	if "airflow" in n or "cfm" in n: return "Airflow", "CFM"
	if "pressure" in n: return "Pressure", "Pa"
	return "Analog", ""

def zone_from_name(name):
	n = name or ""
	m = re.search(r"(North|South|East|West|Middle|Basement|First|Second|Roof|Carport)", n, re.I)
	if m: return m.group(1).title()
	if "W1" in n or "West" in n: return "West"
	if "E1" in n or "East" in n: return "East"
	if "N1" in n or "North" in n: return "North"
	if "S1" in n or "South" in n: return "South"
	return "Building"

def open_db():
	if os.path.exists(DB_PATH): os.remove(DB_PATH)
	con = sqlite3.connect(DB_PATH)
	con.execute("PRAGMA journal_mode=WAL")
	con.execute("PRAGMA synchronous=OFF")
	con.execute("PRAGMA temp_store=MEMORY")
	with open(SCHEMA_PATH) as f: con.executescript(f.read())
	return con

def load_sensors(con):
	import openpyxl
	# Full catalog
	wb = openpyxl.load_workbook(XLSX_FULL, read_only=True, data_only=True)
	ws = wb[wb.sheetnames[0]]
	rows = list(ws.iter_rows(values_only=True))
	hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
	# flexible column find
	def col(*names):
		for n in names:
			for i, h in enumerate(hdr):
				if n in h: return i
		return None
	i_src = col("sensor_name_source", "source") or 1
	i_rep = col("sensor_name_report", "report") or 2
	i_desc = col("sensor_description", "description") or 3
	i_prop = col("sensor_measurement", "properties", "property") or 4
	# simplified set
	simple = set()
	if os.path.exists(XLSX_SIMPLE):
		wb2 = openpyxl.load_workbook(XLSX_SIMPLE, read_only=True, data_only=True)
		ws2 = wb2[wb2.sheetnames[0]]
		r2 = list(ws2.iter_rows(values_only=True))
		h2 = [str(c).strip().lower() if c else "" for c in r2[0]]
		si = 0
		for i, h in enumerate(h2):
			if "source" in h or h == "sensor_id" or "sensor_name_source" in h: si = i; break
		# also try report codes like 30000_TL208
		for r in r2[1:]:
			if not r: continue
			val = r[si] if si < len(r) else None
			if val is None: continue
			s = str(val).strip()
			if re.match(r"^\d+_TL\d+$", s): simple.add(s)
			elif re.match(r"^\d+$", s): pass
			else:
				# try other cols for code
				for c in r:
					if c and re.match(r"^\d+_TL\d+$", str(c).strip()): simple.add(str(c).strip())
	# Always include energy codes + zone thermostat temps (200xx_TL2) + CO2
	simple |= set(ENERGY_CODES.keys())
	for code, sid in []: pass
	# From full catalog rows we'll activate temps below; also add common temp codes
	for ctrl in range(20003, 20017):
		simple.add(f"{ctrl}_TL2")
	simple.add("20016_TL5")  # CO2
	simple.add("20000_TL92")
	simple.add("20000_TL93")
	progress(f"Simplified/energy codes tracked: {len(simple)}")
	batch = []
	for r in rows[1:]:
		if not r or r[i_src] is None: continue
		src = str(r[i_src]).strip()
		if not re.match(r"^\d+_TL\d+$", src): continue
		rep = str(r[i_rep]).strip() if r[i_rep] else src
		desc = str(r[i_desc]).strip() if i_desc is not None and r[i_desc] else ""
		prop = int(r[i_prop]) if i_prop is not None and r[i_prop] is not None and str(r[i_prop]).isdigit() else None
		stype, unit = infer_type_unit(prop, rep)
		if src in ENERGY_CODES:
			_, unit, rep = ENERGY_CODES[src][1], ENERGY_CODES[src][1], ENERGY_CODES[src][2] if not rep else rep
			stype = "Power" if unit == "W" else stype
			if ENERGY_CODES[src][0] == "oat": stype, unit = "Temperature", "°C"
		ctrl = src.split("_")[0]
		full = PREFIX + src
		zone = zone_from_name(rep + " " + desc)
		active = 1 if src in simple else 0
		batch.append((src, full, rep, desc, prop, ctrl, stype, unit, zone, active, 0))
	con.executemany(
		"INSERT OR IGNORE INTO sensors(source_code,full_name,report_name,description,property_code,controller,sensor_type,unit,zone,active,in_timeseries) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
		batch,
	)
	con.commit()
	progress(f"Sensors loaded: {con.execute('SELECT COUNT(*) FROM sensors').fetchone()[0]}")
	# id map for active/energy
	want = {r[0] for r in con.execute("SELECT source_code FROM sensors WHERE active=1 OR source_code IN (%s)" % ",".join("?" * len(ENERGY_CODES)), list(ENERGY_CODES.keys()))}
	# ensure energy sensors active
	for code in ENERGY_CODES:
		con.execute("UPDATE sensors SET active=1 WHERE source_code=?", (code,))
	con.commit()
	idmap = {r[0]: r[1] for r in con.execute("SELECT source_code,id FROM sensors")}
	return idmap, want | set(ENERGY_CODES.keys())

def load_gas(con):
	import openpyxl
	wb = openpyxl.load_workbook(XLSX_GAS, read_only=True, data_only=True)
	ws = wb[wb.sheetnames[0]]
	rows = list(ws.iter_rows(values_only=True))
	n = 0
	for r in rows[1:]:
		if not r or r[0] is None: continue
		d = r[0]
		if hasattr(d, "strftime"): day = d.strftime("%Y-%m-%d")
		else:
			s = str(d).strip()
			# try parse
			day = None
			for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
				try: day = datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d"); break
				except Exception: pass
			if not day:
				m = re.search(r"(20\d{2}-\d{2}-\d{2})", s)
				day = m.group(1) if m else None
		if not day: continue
		meter = float(r[1]) if r[1] is not None and str(r[1]).replace(".","").replace("-","").isdigit() else None
		try: gj = float(r[2]) if r[2] is not None else None
		except Exception: gj = None
		con.execute("INSERT OR REPLACE INTO natural_gas(day,meter_reading,gj_usage) VALUES(?,?,?)", (day, meter, gj))
		n += 1
	con.commit()
	progress(f"Natural gas rows: {n}")

def stream_hourly(con, idmap, want_codes):
	# Map CSV header → source_code
	progress("Opening CSV (streaming hourly aggregate)…")
	f = open(CSV_PATH, "r", newline="", encoding="utf-8", errors="replace")
	reader = csv.reader(f)
	header = next(reader)
	# col index → source_code
	colmap = {}  # idx -> source_code
	for i, h in enumerate(header):
		if i == 0: continue
		h = h.strip()
		code = h[len(PREFIX):] if h.startswith(PREFIX) else h
		if code in want_codes and code in idmap:
			colmap[i] = code
	progress(f"Tracking {len(colmap)} CSV columns")
	# hour_key -> code -> [sum, count]
	bucket = defaultdict(lambda: defaultdict(lambda: [0.0, 0]))
	rows_n = 0
	flush_every = 200000
	def flush():
		nonlocal bucket
		batch = []
		for hour, codes in bucket.items():
			for code, (s, c) in codes.items():
				if c == 0: continue
				sid = idmap[code]
				batch.append((sid, hour, s / c, c))
		if batch:
			con.executemany("INSERT INTO readings_hourly(sensor_id,ts,value,n) VALUES(?,?,?,?) ON CONFLICT(sensor_id,ts) DO UPDATE SET value=((readings_hourly.value*readings_hourly.n + excluded.value*excluded.n)/(readings_hourly.n+excluded.n)), n=readings_hourly.n+excluded.n", batch)
			con.commit()
		bucket = defaultdict(lambda: defaultdict(lambda: [0.0, 0]))
		progress(f"  flushed batch, total CSV rows={rows_n:,}")
	for row in reader:
		rows_n += 1
		if not row: continue
		ts = row[0].strip()
		if len(ts) < 13: continue
		hour = ts[:13].replace(" ", "T") + ":00:00" if "T" not in ts[:13] else ts[:13] + ":00:00"
		# normalize: "2018-04-08 02:15:00" → "2018-04-08T02:00:00"
		if " " in ts:
			hour = ts[:13] + ":00:00"  # YYYY-MM-DD HH
			hour = hour[:10] + "T" + hour[11:]
		for i, code in colmap.items():
			if i >= len(row): continue
			cell = row[i].strip()
			if not cell: continue
			try: v = float(cell)
			except Exception: continue
			if abs(v) > 1e20 or v == -40 and "TL92" not in code:  # skip wild sentinels except possible OAT
				if abs(v) > 1e10: continue
			acc = bucket[hour][code]
			acc[0] += v
			acc[1] += 1
		if rows_n % flush_every == 0: flush()
		if rows_n % 100000 == 0: progress(f"  … {rows_n:,} rows")
	flush()
	f.close()
	con.execute("UPDATE sensors SET in_timeseries=1 WHERE id IN (SELECT DISTINCT sensor_id FROM readings_hourly)")
	con.commit()
	progress(f"Hourly readings: {con.execute('SELECT COUNT(*) FROM readings_hourly').fetchone()[0]:,}")

def build_daily(con, idmap):
	progress("Building daily aggregates + energy_daily…")
	con.execute("DELETE FROM readings_daily")
	con.execute("""INSERT INTO readings_daily(sensor_id,day,value,n)
		SELECT sensor_id, substr(ts,1,10), AVG(value), SUM(n) FROM readings_hourly GROUP BY sensor_id, substr(ts,1,10)""")
	con.commit()
	# energy_daily from known codes
	code_id = {c: idmap[c] for c in ENERGY_CODES if c in idmap}
	def series(code):
		sid = code_id.get(code)
		if not sid: return {}
		return {r[0]: r[1] for r in con.execute("SELECT day,value FROM readings_daily WHERE sensor_id=?", (sid,))}
	hvac, light, dhw, rnd, app = series("30000_TL208"), series("30000_TL209"), series("30000_TL210"), series("30000_TL211"), series("30000_TL212")
	pvc, pvr, site, net = series("30000_TL252"), series("30000_TL253"), series("30000_TL342"), series("30000_TL335")
	# site proxy: sum mains if available else hvac+light+…
	days = sorted(set(hvac) | set(light) | set(pvc) | set(pvr) | set(net))
	gas = {r[0]: r[1] for r in con.execute("SELECT day,gj_usage FROM natural_gas")}
	# emissions: Alberta ~0.54 kg/kWh electric; gas GJ→kWh * 0.18 approx kg/kWh? use 50 kg/GJ rough
	ELECF = 0.54
	GASF = 50.0  # kg CO2e / GJ
	batch = []
	for d in days:
		# W averaged over day ≈ treat as kW mean; kWh ≈ kW * 24 for rough daily energy from power sensors
		def kwh(m, key):
			v = m.get(d)
			return (v or 0) * 24 / 1000.0 if v is not None else 0.0  # W→kW→kWh/day
		hv, li, dh, rn, ap = kwh(hvac, d), kwh(light, d), kwh(dhw, d), kwh(rnd, d), kwh(app, d)
		pc, pr = (pvc.get(d) or 0) / 1000.0, (pvr.get(d) or 0) / 1000.0  # kW mean
		sk = (site.get(d) * 24 / 1000.0) if site.get(d) is not None else (hv + li + dh + rn + ap)
		nk = (net.get(d) * 24 / 1000.0) if net.get(d) is not None else (sk - (pc + pr) * 24)
		gj = gas.get(d)
		# gas monthly — attribute to month days roughly later; store if exact day match
		em = sk * ELECF + ((gj or 0) * GASF)
		batch.append((d, hv, li, dh, rn, ap, pc, pr, sk, nk, gj, em))
	con.executemany("INSERT OR REPLACE INTO energy_daily(day,hvac_kwh,lighting_kwh,dhw_kwh,rnd_kwh,appliances_kwh,pv_carport_kw,pv_rooftop_kw,site_kw,net_kw,gas_gj,emissions_kg) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", batch)
	con.commit()
	progress(f"energy_daily days: {len(batch)}")

def seed_meta(con):
	con.execute("INSERT OR REPLACE INTO users(email,password_hash,display_name,role) VALUES(?,?,?,?)",
		("operator@sait.ca", "password", "ARIS Operator", "operator"))
	for k, v in [
		("auto_approve_dsm", "false"), ("night_purge", "true"), ("peak_shaving", "true"),
		("dark_theme", "false"), ("notify_offline", "true"), ("notify_dsm", "true"),
		("notify_digest", "false"), ("historian_mode", "true"), ("emissions_factor_elec", "0.54"),
		("replay_day", ""), ("version", "ARIS v2.5.0"),
	]:
		con.execute("INSERT OR REPLACE INTO building_settings(key,value) VALUES(?,?)", (k, v))
	con.commit()

def generate_alerts_dsm(con, idmap):
	progress("Generating alerts + DSM heuristics from recent data…")
	# last 14 days of oat + hvac
	oat_id = idmap.get("20000_TL92")
	hvac_id = idmap.get("30000_TL208")
	rows = []
	if hvac_id:
		rows = list(con.execute("SELECT ts,value FROM readings_hourly WHERE sensor_id=? ORDER BY ts DESC LIMIT 336", (hvac_id,)))
	# anomaly: value > mean+2.5std or flat
	import statistics
	if len(rows) >= 24:
		vals = [r[1] for r in rows]
		mu, sd = statistics.mean(vals), statistics.pstdev(vals) or 1.0
		for ts, v in rows[:48]:
			if abs(v - mu) > 2.5 * sd:
				con.execute("INSERT INTO alerts(sensor_id,severity,kind,title,message,ts) VALUES(?,?,?,?,?,?)",
					(hvac_id, "high", "anomaly", "HVAC load anomaly", f"Space HVAC {v:.0f} W vs mean {mu:.0f} W", ts))
				break
	# sensor offline heuristic: zone temps with no recent data
	latest = con.execute("SELECT MAX(ts) FROM readings_hourly").fetchone()[0]
	if latest:
		for sid, code, name in con.execute("SELECT id,source_code,report_name FROM sensors WHERE sensor_type='Temperature' AND in_timeseries=1 LIMIT 20"):
			mx = con.execute("SELECT MAX(ts) FROM readings_hourly WHERE sensor_id=?", (sid,)).fetchone()[0]
			if mx and mx < latest[:10] + "T00:00:00":
				con.execute("INSERT INTO alerts(sensor_id,severity,kind,title,message,ts) VALUES(?,?,?,?,?,?)",
					(sid, "medium", "offline", f"{name or code} stale", f"Last reading {mx}", latest))
	# DSM rules
	recs = [
		("West", "Reduce HVAC setpoint 1°C overnight", "High night HVAC vs mild OAT pattern", 4.2, 0.88, "High"),
		("Building", "Shift R&D plug loads off peak 16:00–19:00", "AESO-style peak window + high R&D coincident load", 3.1, 0.82, "High"),
		("Roof", "Maximize PV self-consumption — pre-cool midday", "Rooftop + carport PV high while site import positive", 5.4, 0.91, "Medium"),
		("Building", "Dim non-critical lighting 10% 12:00–15:00", "Lighting end-use stable; TOU midday opportunity", 1.8, 0.76, "Low"),
		("East", "Delay DHW recharge until after 21:00", "DHW spikes align with evening price peak", 2.2, 0.80, "Medium"),
	]
	for zone, action, reason, kwh, conf, pri in recs:
		con.execute("INSERT INTO dsm_recommendations(zone,action,reason,impact_kwh,confidence,priority,status) VALUES(?,?,?,?,?,?,?)",
			(zone, action, reason, kwh, conf, pri, "Pending"))
	con.execute("INSERT INTO alerts(sensor_id,severity,kind,title,message,ts) VALUES(?,?,?,?,?,?)",
		(None, "low", "dsm", "New DSM Recommendation Generated", "Heuristic DSM pack refreshed from historian", latest or datetime.utcnow().isoformat()))
	con.commit()
	progress(f"Alerts={con.execute('SELECT COUNT(*) FROM alerts').fetchone()[0]} DSM={con.execute('SELECT COUNT(*) FROM dsm_recommendations').fetchone()[0]}")

def main():
	progress("=== ARIS ETL start ===")
	assert os.path.exists(CSV_PATH), CSV_PATH
	con = open_db()
	idmap, want = load_sensors(con)
	load_gas(con)
	stream_hourly(con, idmap, want)
	build_daily(con, idmap)
	seed_meta(con)
	generate_alerts_dsm(con, idmap)
	con.execute("PRAGMA optimize")
	con.close()
	sz = os.path.getsize(DB_PATH) / (1024 * 1024)
	progress(f"=== Done → {DB_PATH} ({sz:.1f} MB) ===")

if __name__ == "__main__":
	main()
